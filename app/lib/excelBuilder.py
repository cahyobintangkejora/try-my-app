from decimal import Decimal
from typing import Dict
from xlsxwriter import Workbook
from io import BytesIO
from flask import send_file, Response
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
from flask import __version__ as flask_version

'''
ExcelBuilder
Versi: 3.8 (03 Feb 2023)
'''


class ExcelBuilder():
    ''' Develop by Candra 22-06-2022
        Contoh pemakaian ada di file contohExcelBuilder.py
        Pada Constructor class ini menerima 6 parameter:
            1. dataHeader(wajib): data berupa header yg nanti di-cetak dari atas ke bawah ['PT. SAT', 'CABANG: KZ01', ..]
            2. tableHeader(wajib): ada 2 tipe, yg normal ['NO', 'NAMA'], dan nested ['NO', ['HARI', 'SENIN', 'SELASA'], 'TOTAL']
            3. sheetName(opsional): untuk nama worksheet, default 'Sheet1'
            4. dateTimeFormat(optional): untuk list format waktu dalam python format, untuk format rata kanan
            5. headerDefaultFormat(optional): untuk default cell format pada header excel
            6. theadDefaultFormat(optional): untuk default cell format pada table header
    '''

    def __init__(self, dataHeader: list, tableHeader: list, sheetName='Sheet1', dateTimeFormat: list = [], headerDefaultFormat='format_0', theadDefaultFormat='format_3') -> None:
        # wajib pakai flask versi >=2
        if int(flask_version[0]) < 2:
            raise Exception(
                f"Maaf, excel builder memerlukan flask versi 2 keatas. Versi flask saat ini: {flask_version}"
            )

        # private variable
        self.__ouputFile = BytesIO()
        self.__wb = Workbook(self.__ouputFile, {'in_memory': True})
        self.__ws = self.__wb.add_worksheet(sheetName)
        self.__row_count = 2
        self.__sumCols = {}
        self.__tableHeader = []
        self.__flagBody = False
        self.__flagRight = False
        self.__commit = True
        self.__last_col = 0
        self.__dataRowTemp = []
        self.__listDateTimeFormat = [
            '%d %b %Y',
            '%d-%b-%Y',
            '%d/%b/%Y',
            '%d %m %Y'
            '%d-%m-%Y',
            '%d/%m/%Y',
            '%Y %m %d',
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%d %mon',
            '%Y/%m%d',
            '%Y-%m-%d'
        ]
        self.__listDateTimeFormat = dateTimeFormat if dateTimeFormat else self.__listDateTimeFormat
        self.__theadDefaultFormat = theadDefaultFormat
        self.__headerDefaultFormat = headerDefaultFormat

        # call private method to setup excel file
        self.__createCellFormats()
        self.__defineTableHeader(tableHeader)
        self.__setHeader(dataHeader)
        self.__setTableHeader(tableHeader)
        self.__defineMaxCharCols()

    def __createCellFormats(self) -> None:
        # no format
        self.format_0 = self.__wb.add_format({})
        # all border
        format_1 = {
            'border': 2,
        }
        self.format_1 = self.__wb.add_format(format_1)
        # all border & text bold
        format_2 = {
            'border': 2,
            'bold': True
        }
        self.format_2 = self.__wb.add_format(format_2)
        # all border,text bold & align center, vcenter
        format_3 = {
            'border': 2,
            'bold': True,
            'align': 'center',
        }
        self.format_3 = self.__wb.add_format(format_3)
        self.format_3.set_align('vcenter')
        # list cell_format dalam object
        self.__listFormat = {
            'format_0': self.format_0,
            'format_1': self.format_1,
            'format_2': self.format_2,
            'format_3': self.format_3
        }
        # list cell_format dalam dict (raw)
        self.__listFormatRaw = {
            'format_0': {},
            'format_1': format_1,
            'format_2': format_2,
            'format_3': format_3
        }

    def __get_huruf(self, index) -> str:
        return get_column_letter(index + 1)

    def __getIndexHuruf(self, huruf) -> int:
        index_huruf = column_index_from_string(huruf)
        return index_huruf + 1

    def __defineTableHeader(self, tableHeader) -> None:
        for header in tableHeader:
            if isinstance(header, list):
                # untuk kolom yg memiliki cabang/anak kita perlu decode __tableHeader,
                # agar hanya menggunakan kolom cabang/anak-nya saja,
                # ex: ['NO', ['HARI', 'SENIN', 'SELASA'], ...], kita ambil 'NO', 'SENIN', 'SELASA' saja (cabang/anak-nya)
                for j in header[1:]:
                    if isinstance(j, tuple):
                        j = j[0]

                    self.__tableHeader.append(j.upper())
            else:
                if isinstance(header, tuple):
                    header = header[0]
                self.__tableHeader.append(header.upper())

    def __defineMaxCharCols(self) -> None:
        # method untuk membuat variabel penampung max char per kolom
        self.__maxCharCols = {}
        for count, value in enumerate(self.__tableHeader):
            # nilai default panjang col adalah panjang dari table header-nya,
            # misal, thead = 'NAMA' maka default width untuk col 'NAMA' adalah 4
            self.__maxCharCols[count] = len(value)

    def __computeColsWidth(self, dataRow) -> None:
        # method ini digunakan untuk hitung ulang lebar minimum suatu kolom
        for index, val in enumerate(dataRow):
            # get len_val
            if isinstance(val, tuple):
                val = '' if val[0] is None else str(val[0])
                len_val = len(val)
            else:
                val = '' if val is None else str(val)
                len_val = len(val)

            # update maxCharCols
            if self.__maxCharCols.get(index) is None:
                self.__maxCharCols[index] = 0
            if len_val > self.__maxCharCols[index]:
                self.__maxCharCols[index] = len_val

    def __setColsWidth(self) -> None:
        # jika table header adalah NO
        if self.__tableHeader[0][0:2] == 'NO':
            self.__maxCharCols[0] = 0
        for count in range(len(self.__maxCharCols)):
            # set lebar col berdasarkan panjang char tiap cell ditambah 5
            width = self.__maxCharCols[count] + 5
            # proteksi, panjang maksimal setelah di-tambah 5 adalah 50
            width = 50 if width > 50 else width
            col = f"{self.__get_huruf(count)}:{self.__get_huruf(count)}"
            self.__ws.set_column(col, width=width)

    def __setHeader(self, dataHeader) -> None:
        lenCol = len(self.__tableHeader) - 1
        for data in dataHeader:
            # decodeCellData return tuple -> ('TANGGAL', 'A', 6, <xlsxwriter.format.Format object at 0x25A>)
            # kita ambil cell value & cell format-nya saja
            cell_value, *others, cell_format = self.__decodeCellData(
                cell_data=data,
                defaultFormat=self.__headerDefaultFormat
            )

            cell_merge = f"A{self.__row_count}:{self.__get_huruf(lenCol)}{self.__row_count}"
            self.__ws.merge_range(cell_merge, cell_value, cell_format)
            self.__row_count += 1
        # create row break
        self.__ws.merge_range(
            f"A{self.__row_count}:{self.__get_huruf(lenCol)}{self.__row_count}", None
        )
        self.__row_count += 1

    def __setTableHeader(self, tableHeader) -> None:
        '''
            Method ini untuk membuat table header (thead) pada excel,
            thead memiliki 2 tipe, yg biasa dan yg memiliki anak/cabang,
            eg: thead normal-> ['NO', 'TOTAL'], thead dgn cabang/anak-> ['NO', ['HARI', 'SENIN', 'SELASA'], 'TOTAL']
            jika thead normal semua type-nya str, sedangkan untuk thead yg ada cabang/anak ditulis dalam list,
            index 0 adalah root/parent dan index selanjutnya adalah cabang/child
        '''
        last_col = 0
        if all([type(i) in {str, tuple} for i in tableHeader]):
            for count, value in enumerate(tableHeader):
                # thead bisa disetting format cell-nya dengan cara taruh di tuple, contoh ('NAMA', {'font_color': 'red}),
                # sehingga perlu kita decode jika value merupakan tuple
                cell_value, *others, cell_format = self.__decodeCellData(
                    cell_data=value,
                    defaultFormat=self.__theadDefaultFormat
                )

                # row_count - 1 karena row_count nilai awalnya adalah 2,
                # kenapa nilai awalnya 2??, itu aku ngikut yg ada di-oracle form-nya (bisa diganti jika tidak cocok)
                self.__ws.write(
                    self.__row_count - 1, count, cell_value, cell_format
                )
            self.__row_count += 1
        else:
            # tableHeader -> ['NO', ['HARI', 'SENIN', 'SELASA]]
            for count, value in enumerate(tableHeader):
                if isinstance(value, list) and len(value) > 1:
                    # jika value adalah list -> ['HARI', 'SENIN', 'SELASA'],
                    # maka value[0] -> ['HARI'] sebagai parent thead,
                    # dan value[1:] -> ['SENIN', 'SELASA'] adalah child thead
                    parent_thead = value[0]
                    child_thead = value[1:]

                    # ===== parent thead ===== #
                    # decodeCellData return tuple -> ('TANGGAL', 'A', 6, <xlsxwriter.format.Format object at 0x25A>)
                    # kita ambil cell value & cell format-nya saja
                    cell_value, *others, cell_format = self.__decodeCellData(
                        cell_data=parent_thead,
                        defaultFormat=self.__theadDefaultFormat
                    )

                    cell_start = f'{self.__get_huruf(last_col)}{self.__row_count}'
                    cell_end = f'{self.__get_huruf(last_col + (len(value) - 2))}{self.__row_count}'
                    if cell_start == cell_end:
                        self.__ws.write(cell_start, cell_value, cell_format)
                    else:
                        self.__ws.merge_range(
                            f'{cell_start}:{cell_end}', cell_value, cell_format
                        )

                    # ===== child thead ===== #
                    for i in child_thead:
                        # decodeCellData return tuple -> ('TANGGAL', 'A', 6, <xlsxwriter.format.Format object at 0x25A>)
                        # kita ambil cell value & cell format-nya saja
                        cell_value, *others, cell_format = self.__decodeCellData(
                            cell_data=i,
                            defaultFormat=self.__theadDefaultFormat
                        )

                        cell_start = f'{self.__get_huruf(last_col)}{self.__row_count + 1}'
                        self.__ws.write(
                            cell_start, cell_value, cell_format
                        )
                        last_col += 1

                else:
                    # decodeCellData return tuple -> ('TANGGAL', 'A', 6, <xlsxwriter.format.Format object at 0x25A>)
                    # kita ambil cell value & cell format-nya saja
                    cell_value, *others, cell_format = self.__decodeCellData(
                        cell_data=value,
                        defaultFormat=self.__theadDefaultFormat
                    )

                    cell_start = f'{self.__get_huruf(last_col)}{self.__row_count}'
                    cell_end = f'{self.__get_huruf(last_col)}{self.__row_count + 1}'
                    self.__ws.merge_range(
                        f'{cell_start}:{cell_end}', cell_value, cell_format
                    )
                    last_col += 1
            self.__row_count += 2

    def __getCellFormat(self, defaultFormat: dict) -> dict:
        '''
            Method ini untuk memodifikasi cell_format agar menjadi rata kanan jika cell_value adalah format rupiah atu tanggal (aturan atasan).
            defaultFormat di-isi oleh dict yang bisa diparse ke xlsxwriter.format.Format
            Return method ini adalah dict yg nantinya akan di-decode menjadi xlsxwriter.format.Format oleh method decodeCellFormat()
        '''

        # jika flagRight True maka itu adalah kolom angka(rupiah) atau tanggal,
        # sehingga harus dibuat format rata kanan (aturan dari atasan), kecuali disetting berbeda
        if self.__flagRight and defaultFormat:
            cell_format = self.__decodeCellFormat(
                cell_format=defaultFormat,
                toDict=True
            )
            cell_format['align'] = 'right'
        elif self.__flagRight:
            cell_format = {
                'align': 'right'
            }
        else:
            cell_format = defaultFormat
        self.__flagRight = False
        return cell_format

    def __decodeCellFormat(self, cell_format, defaultFormat=None, toDict=False):
        '''
            Method ini menerima 3 argumen:
                a. cell_format: merupakan format yg hendak di decode. contoh: 'format_3' atau {'bold': True}
                b. defaultFormat: jika cell_format gagal di-decode / tidak diberikan format, maka gunakan defautFormat. contoh: {'font_color': 'red'} atau 'format_3'
                c. toDict, jika True maka return cell_format dalam bentuk dict alih-alih return dalam bentuk xlsxwriter.format.Format obj jika False. contoh: True atau False
            Argumen pada param defaultFormat atau cell_format yang di-inputkan ada 2 tipe:
                a. dict: {'font_color': 'red}
                b. str: 'format_3' -> yg akan di translate jadi -> {'border': 2, 'bold': True, 'align': 'vcenter'}
            Output dari method ini ada 2:
                a. jika toDict=False -> xlsxwriter.format.Format,
                    contoh: decodeCellFormat('format_3', toDict=False) nanti outputnya <xlsxwriter.format.Format object at 0x25A>
                b. jika toDict=True -> dict,
                    contoh: decodeCellFormat('format_3', toDict=True) nanti outputnya {'border': 2, 'bold': True, 'align': 'vcenter'}
            Alasan perlu return dalam bentuk dict adalah untuk bisa diformat lagi ditambahkan format rata kanan pada method getCellFormat().
        '''
        v_defaultFormat = defaultFormat if not defaultFormat else self.format_0
        # jika user bikin format sendiri, maka kita buatkan formatnya
        # param toDict digunakan untuk return cell_format dalam bentuk dict jika True
        if isinstance(cell_format, dict):
            return cell_format if toDict else self.__wb.add_format(cell_format)
        elif isinstance(cell_format, str):
            # jika format yg diberikan user tidak ada maka gunakan default format
            return self.__listFormatRaw.get(cell_format, {}) if toDict else self.__listFormat.get(cell_format, v_defaultFormat)
        else:
            return cell_format

    def __decodeCellValue(self, cell_value):
        '''
            Method ini untuk decode cell_value, ada 3 tipe decode:
                1. jika tipe str maka cek apakah user ingin ambil sumCols? (eg: 'SUM PPN' / 'GRAND_PPN')
                2. jika tipe Decimal/float maka ubah jadi rupiah (eg: 1500.255 -> 1.500,25)
                3. jika tipe datetime kita ubah ke string tanggal -> 01/01/2022 
            Method ini juga menandai flagRight jika tipenya adalah float/Decimal/tanggal,
            yg nantinya akan diformat rata kanan(aturan atasan) pada method getCellFormat().
        '''
        # decode cell_value dulu jika cell_value berupa string
        # berfungsi untuk cek apakah ambil sumCols? jika ya maka return sumCols,
        # atau jika string bisa di-parse ke tanggal, sehingga flagRight True (untuk format rata kanan)
        cell_value = self.getSum(cell_value)
        if type(cell_value) in {float, Decimal}:
            cell_value = f"{cell_value:,.2f}".translate(
                str.maketrans(',.', '.,')
            )
            self.__flagRight = True
        elif isinstance(cell_value, datetime):
            cell_value = cell_value.strftime(self.__listDateTimeFormat[0])
            self.__flagRight = True
        return cell_value

    def __decodeEndCol(self, data) -> str:
        v_last_col = self.__get_huruf(self.__last_col)
        if isinstance(data, tuple) and len(data) > 1:
            if type(data[1]) not in {dict, int} and len(data[1]) < 5:
                v_last_col = data[1]
        return v_last_col

    def __decodeEndRow(self, data) -> int:
        # method ini masih dalam tahap pengembangan untuk merge row kebawah nantinya!
        v_last_row = self.__row_count
        if isinstance(data, tuple) and len(data) > 1 and isinstance(data[1], int):
            v_last_row += data[1]
            # karena masih belum boleh dipakai, kita kasih error jika ada yg maksa pakai :)
            raise Exception(
                "Maaf, fitur merge cell kebawah masih dalam tahap pengembangan!"
            )
        return v_last_row

    def __decodeCellData(self, cell_data, defaultFormat=None) -> tuple:
        '''
            cell_data itu merupakan meta-data dari cell, didalamnya ada 4 atribut:
                a. cell_value: merupakan value yang akan ditulis / tampilkan di cell excel.
                b. end_col: merupakan col dimana cell akan ditulis.
                c. end_row: merupakan row dimana cell akan ditulis.
                d. cell_format: mengatur tampilan dan format dari cell.
            Method ini dibuat untuk membuat meta-data dari cell. Menerima 2 argumen:
            1. cell_data(wajib): merupakan cell meta-data, ada 2 bentuk:
                a.  normal / raw data / cell_value (str/int/dll), contoh: 'DC CIKOKOL', 2025, Decimal(123xxx).
                b.  tuple: jika mau ubah tampilan cell, merge cell ke kanan, bisa ditaruh dalam tuple,
                    contoh tuple: ('DC CIKOKOL', 'C', {'font_color': 'red}) artinya: tulis cell_value 'DC CIKOKOL' dgn warna merah, dan juga merge cell ke kanan hingga col 'C'.
            2. defaultFormat (opsional): merupakan format tampilan dasar cell excel, secara default adalah None artinya tidak ada format sama sekali / polosan.
            Return dari Method ini adalah tuple dgn format: (cell_value, end_col, end_row, cell_format). Contoh:
                a. INPUT: __decodeCellData(('TANGGAL', {'font_color': 'red}))
                b. OUTPUT: ('TANGGAL', 'A', 6, <xlsxwriter.format.Format object at 0x25A>)
        '''

        if isinstance(cell_data, tuple):
            '''
                Jika cell_data berupa tuple. format tuple ada 3:
                    1. (CELL_VALUE, END_COL, CELL_FORMAT)
                    2. (CELL_VALUE, END_COL)
                    3. (CELL_VALUE, CELL_FORMAT)
                Maka decode dulu cell_data-nya untuk mendapatkan cell_value, end_col (jika ada), end_row, dan cell_format (jika ada)
            '''
            # index ke-0 (cell_data[0]) sudah pasti adalah cell_value yg akan ditulis ke cell excel
            # contoh: ('candra', 'H', {'bold': True}) atau ('candra', 'H') atau ('candra', {'bold': True}) -> index-0 pasti 'candra' atau cell_value-nya
            r_cell_value = self.__decodeCellValue(cell_data[0])
            r_end_col = self.__decodeEndCol(cell_data)
            r_end_row = self.__decodeEndRow(cell_data)
            cell_format = self.__getCellFormat(defaultFormat)

            # jika user provide format, maka ambil format yg user berikan
            # format wajib ditaruh pada index ke-3 atau ke-2. sehingga cek apakah len > 2
            if len(cell_data) > 2:
                # pada index ke-2 (cell_data[2]) sudah pasti isinya adalah format yg user berikan
                # contoh: ('candra', 'H', {'bold': True}) -> index-3 pasti {'bold': True} atau cell_format-nya
                cell_format = cell_data[2]
            elif len(cell_data) > 1:
                # pada index ke-1 (cell_data[1]) itu ada dua kemungkinan, bisa jadi end_col atau cell_format,
                # untuk itu kita cek apakah index 1 tipenya dict (eg: {'font_color': 'red'}) atau panjang-nya > 5 (eg: 'format_3') ?
                if isinstance(cell_data[1], dict) or len(cell_data[1]) > 5:
                    cell_format = cell_data[1]

            r_cell_format = self.__decodeCellFormat(
                cell_format,
                defaultFormat=defaultFormat
            )

            # r -> result -> r_cell_value -> result_cell_value
            # contoh return -> ('TANGGAL', 'A', 6, <xlsxwriter.format.Format object at 0x25A>)
            return r_cell_value, r_end_col, r_end_row, r_cell_format
        else:
            r_cell_value = self.__decodeCellValue(cell_data)
            r_end_col = self.__get_huruf(self.__last_col)
            r_end_row = self.__decodeEndRow(cell_data)
            r_cell_format = self.__decodeCellFormat(
                cell_format=self.__getCellFormat(defaultFormat),
                defaultFormat=defaultFormat
            )

            # r -> result -> r_cell_value -> result_cell_value
            # contoh return -> ('TANGGAL', 'A', 6, <xlsxwriter.format.Format object at 0x25A>)
            return r_cell_value, r_end_col, r_end_row, r_cell_format

    def __insertRow(self, dataRow: list, defaultFormat=None, commit=True) -> None:
        # jika commit False, berarti masih melanjutkan dari lastCol terakhir,
        # jika commit True, berarti kita mulai dari 0 ya (kaya pertamina)
        self.__last_col = 0 if self.__commit else self.__last_col

        # ngoding itu sulit ya :')
        # v_dataRow digunakan untuk menampung dataRow yg akan di-hitung panjang char-nya
        v_dataRow = []
        for data in dataRow:
            cell_value, end_col, end_row, cell_format = self.__decodeCellData(
                data, defaultFormat
            )
            if self.__get_huruf(self.__last_col) == end_col and end_row == self.__row_count:
                self.__ws.write(
                    f"{self.__get_huruf(self.__last_col)}{self.__row_count}", cell_value, cell_format
                )
                v_dataRow.append(cell_value)
            else:
                self.__ws.merge_range(
                    f"{self.__get_huruf(self.__last_col)}{self.__row_count}:{end_col}{end_row}", cell_value, cell_format
                )
                # untuk cell yg merge tidak dihitung ulang panjang char tiap cell, sehingga di-isi oleh ''
                v_dataRow.extend(
                    ['' for i in range(self.__getIndexHuruf(end_col) - 1)]
                )
            self.__last_col = self.__getIndexHuruf(end_col) - 1

        # tiap kali insert row maka wajib hitung ulang panjang char tiap cell,
        # ini bertujuan untuk menentukan panjang/width tiap kolom (auto fit cells width)
        if not commit:
            self.__dataRowTemp.extend(v_dataRow)
        elif commit:
            if len(self.__dataRowTemp) > 0:
                self.__dataRowTemp.extend(v_dataRow)
                self.__computeColsWidth(self.__dataRowTemp)
                self.__dataRowTemp.clear()
            else:
                self.__computeColsWidth(v_dataRow)
        else:
            self.__dataRowTemp.clear()

        # insert row sudah selesai, maka tak lupa tambah row_count
        if commit:
            self.__row_count += 1

    def commitRow(self) -> None:
        '''
            Method untuk commit row, jadi mulai di row dan col baru dibawahnya,
            sehingga pemanggilan insertRow() atau insertBody() berikutnya akan dimulai di row berikutnya.
            Ilustrasi: kalau insertRow commit False itu seperti print di terminal tapi nggak ada \n (enter),
                nah kalau mau enter kan print('\n'), commitRow() ini seperti itu juga gunanya.
        '''
        self.__computeColsWidth(self.__dataRowTemp)
        self.__dataRowTemp.clear()
        self.__row_count += 1
        self.__last_col = 0

    def getSum(self, data) -> Decimal:
        '''
            Method getSum digunakan untuk ambil SUM dari suatu col. contoh: data('SUM PPN') maka ambil sum/total dari kolom 'PPN'.
            Method ini juga untuk menentukan apakah data (cell_value) itu nanti ditulis dengan format rata kanan atau tidak.
        '''
        if isinstance(data, str) and data[0:4].upper() == 'SUM ':
            self.__flagRight = True
            if self.__flagBody is True:
                data_key = data.split(' ')[1:]
                data_key = ' '.join(map(str, data_key)).upper()
                data_value = self.__sumCols.get(data_key)
                data = data_value
                # proteksi, jika kolom SUM yg diminta tidak ada
                if data is None:
                    raise Exception(
                        f"Maaf, kolom '{data_key}' tidak ada! Berikut kolom yang ada: {self.__sumCols}"
                    )
                # compute grand total
                if self.__sumCols.get(f'GRAND_{data_key}') is None:
                    self.__sumCols[f'GRAND_{data_key}'] = Decimal(0)
                self.__sumCols[f'GRAND_{data_key}'] += Decimal(data_value)
            elif self.__flagBody is False:
                data = 0.0
        elif isinstance(data, str):
            # jika data memungkinkan untuk di-parse ke datetime maka flagRight True
            # note: jangan diganti pakai dateutil.parser.parse ya walaupun lebih simpel :)
            for fmt in self.__listDateTimeFormat:
                try:
                    datetime.strptime(data, fmt)
                    self.__flagRight = True
                    break
                except Exception:
                    self.__flagRight = False
        return data

    def insertBody(self, dataBody: list, format: Dict = {}) -> None:
        if dataBody is None:
            return
        # tiap kali insert body sumCols dihitung ulang
        for k, v in self.__sumCols.items():
            if k.upper()[0:6] != 'GRAND_':
                self.__sumCols[k] = Decimal(0)

        for count, data in enumerate(dataBody, 1):
            if len(data) == 0:
                return

            v_data = {}
            for k, v in data.items():
                # make key to upper case
                k = k.upper()
                v_data[k] = v

                # isi sumCols
                if type(v) in {Decimal, float}:
                    try:
                        # compute normal sum
                        if self.__sumCols.get(k) is None:
                            self.__sumCols[k] = Decimal(0)
                        self.__sumCols[k] += Decimal(v)
                        # define grand sum
                        if self.__sumCols.get(f'GRAND_{k}') is None:
                            self.__sumCols[f'GRAND_{k}'] = Decimal(0)
                    except Exception as e:
                        pesan = f'error di compute normal sum: {e}'
                        raise Exception(pesan)
            data = v_data

            # decode format
            if format:
                for k, v in format.items():
                    # key adalah kolom yg ingin diformat. ex: 'DPP', 'PPN', 'TOTAL
                    key = k.upper()
                    # func merupakan fungsi dgn return cell_format,
                    # eg: lambda x: {'font_color': 'red'} if x < 10 else {'font_color': 'blue'}
                    #   -> artinya jika nilai pada kolom n < 10 maka warnanya merah, jika >=10 jadi warna biru
                    #   -> param x adalah nilai dari kolom yg ingin di format (key)
                    func = v

                    # decode key, misal 'DPP' itu ada di col 'I'
                    col = self.__get_huruf(self.__tableHeader.index(key))
                    # decode format
                    fmt = func(data[key])

                    data[key] = (data[key], col, fmt)

            # decode dataRow
            dataRow = []
            if self.__tableHeader[0].upper()[0:2] == 'NO':
                dataRow = [
                    data.get(i.upper(), '') for i in self.__tableHeader[1:]
                ]
                dataRow.insert(0, (count, {'align': 'left'}))
            else:
                dataRow = [data.get(i.upper(), '') for i in self.__tableHeader]

            # begin insert row
            self.__insertRow(dataRow)
            self.__flagBody = True

    def insertRow(self, dataRow, defaultFormat=None, commit: bool = True) -> None:
        # default format pada insert row adalah format_0 yaitu tidak ada format (polosan)
        if type(dataRow) is not list and dataRow is not None:
            # jika dataRow yg diinput user adalah dict maka perlu kita susun urutan insert-nya dalam list
            if isinstance(dataRow, dict):
                data = {k.upper(): v for k, v in dataRow.items()}
                dataRow = [
                    data.get(i.upper(), '') for i in self.__tableHeader
                ]
            # jika dataRow yg diinput user bukan dict, maka masukan dalam list
            else:
                dataRow = [dataRow]

        if type(dataRow) in {list, tuple} and len(dataRow) < 1:
            return

        self.__insertRow(
            dataRow=dataRow, defaultFormat=defaultFormat, commit=commit
        )
        self.__commit = commit

    def createExcel(self, fileName) -> Response:
        # method ini untuk buat file excel dalam bentuk attachment flask,
        # tinggal di return di controller.
        self.__setColsWidth()
        self.__wb.close()
        self.__ouputFile.seek(0)

        return send_file(self.__ouputFile, mimetype="application/vnd.ms-excel", as_attachment=True, download_name=f'{fileName}.xlsx', max_age=-1)

    def saveToFile(self, fileName) -> None:
        # method ini untuk save file excel ke direktori lokal,
        # hanya digunakan untuk demo/coba2 aja biasanya
        self.__setColsWidth()
        self.__wb.close()
        self.__ouputFile.seek(0)

        with open(f'{fileName}.xlsx', 'wb') as outfile:
            outfile.write(self.__ouputFile.getbuffer())
